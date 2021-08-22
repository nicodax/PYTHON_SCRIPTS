#!/usr/bin/env python 3

# Imports
import openpyxl, json
from openpyxl import Workbook

# JSON file name conventions
json_file_number = 1
json_file_name = "file" + str(json_file_number)

# Functions
def getHeaders(json_data):
    """
    Get all keys from specified dictionary

    Parameters
    ----------
        json_data : list (of dictionaries)
            data to get keys from
    
    Returns
    ----------
        headers : list
            all specified dictionary's keys
    """
    headers = json_data[0].keys()
    return headers

def setheaderRow(headers, ws):
    """
    Set "header row" to specified worksheet's specified row

    Parameters
    ----------
        headers : list
            list of all categories
        ws : openpyxl worksheet
            openpyxl worksheet to write to
    """
    col = 1
    row = 1
    for header in headers:
        ws.cell(row, col, header)
        col += 1

def setValues(json_data, headers, ws):
    """
    Set values from specified data to specified worksheet's

    Parameters
    ----------
        json_data : list (of dictionaries)
            data to write to openpyxl worksheet
        headers : list
            list of all categories
        ws : openpyxl worksheet
            openpyxl worksheet to write to
    
    Returns
    ----------
        different_json_data : list (of dictionaries)
            all json data not matching the first keys pattern
    """
    col = 1
    row = 2
    different_json_data = []

    for json_object in json_data:
        if (json_object.keys() == headers):
            for category in json_object:
                # Transform lists to strings
                if isinstance(json_object[category], list):
                    json_object[category] = ', '.join([str(item) for item in json_object[category]])             
                ws.cell(row, col, json_object[category])
                col += 1
            row += 1
            col = 1
        else:
            different_json_data.append(json_object)

    return different_json_data

if __name__ == "__main__":
    # Get JSON data
    json_data = {}
    with open(json_file_name + '.json') as json_file:
        json_data = json.load(json_file)
    
    # Initialize openpyxl Workbook
    wb = Workbook()
    ws_1 = wb.active
    ws_1.title = "JXCEL SHEET 1"

    # Get headers
    headers = getHeaders(json_data)

    # Set header row
    setheaderRow(headers, ws_1)

    # Set values
    different_json_data = setValues(json_data, headers, ws_1)
    
    sheetNumber = 2

    while (different_json_data):
        # Initialise new openpyxl sheet
        dynamicSheetName = "ws_" + str(sheetNumber)
        globals()[dynamicSheetName] = wb.create_sheet("sheet" + str(sheetNumber))
        globals()[dynamicSheetName].title = "JXCEL SHEET " + str(sheetNumber)
        sheetNumber += 1

        # Get headers
        headers = getHeaders(different_json_data)

        # Set header row
        setheaderRow(headers, globals()[dynamicSheetName])
        
        # Set values
        different_json_data = setValues(different_json_data, headers, globals()[dynamicSheetName])

    # Save excel file
    wb.save(json_file_name + ".xlsx")
